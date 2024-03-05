from selenium import webdriver as opcSelenium
from selenium.webdriver.common.by import By
import openpyxl

url = "https://www.infomoney.com.br/ferramentas/cambio/"

navegador = opcSelenium.Edge()
navegador.get(url)

workbook = openpyxl.Workbook()
sheet = workbook.active

elemento_tabela = navegador.find_element(By.XPATH, '//*[@id="container_table"]/table')

linhas = elemento_tabela.find_elements(By.TAG_NAME, "tr")

for linhaAtual, linha in enumerate(linhas, start=1):

    cabecalho_tabela = linha.find_elements(By.TAG_NAME, "th")

    celulas = linha.find_elements(By.TAG_NAME, "td")

    if cabecalho_tabela:
        for cabecalho_linha, cabecalho in enumerate(cabecalho_tabela, start=1):

            sheet.cell(row=cabecalho_linha, column=linhaAtual, value=cabecalho.text)
    else:
        for celulaAtual, celula in enumerate(celulas, start=1):

            sheet.cell(row=celulaAtual, column=linhaAtual, value=celula.text)

workbook.save("dados_tabela.xlsx")

navegador.quit()